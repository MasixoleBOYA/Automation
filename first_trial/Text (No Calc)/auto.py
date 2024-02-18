from openpyxl import Workbook, load_workbook
import os

current_path = os.getcwd()

print(f"--CURRENT WORKING DIRECTORY--: \n{current_path}\n")
# importing the worksheets

wb_student_names = load_workbook('first_trial/Text (No Calc)/mock1_StudentNames.xlsx') # will copy from this
destination_sheet = wb_student_names['Sheet1']

wb_student_ages = load_workbook('first_trial/Text (No Calc)/mock2_StudentAges.xlsx') # will copy into this
source_sheet = wb_student_ages['Sheet1']

ages = []
for row in source_sheet.iter_rows(values_only=True):
    for cell in row:
        ages.append(cell.value)

for index,age in enumerate(ages, start = 2):
    destination_sheet.cell(column =3, row=index, value=age)

wb_student_names.save('first_trial/Text (No Calc)/mock1_StudentNames_UPDATED.xlsx')
